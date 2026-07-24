"""Bulk Processor for validating addresses from uploaded Excel files.

Reads .xlsx/.xls files, extracts address rows, validates each against the
USPS adapter via AddressValidator, and returns ordered results.
"""

from typing import List

from fastapi import UploadFile
from openpyxl import load_workbook

from app.models.address import StructuredAddress
from app.models.validation import BulkValidationResult, ValidationResult
from app.services.address_validator import AddressValidator


class InvalidFileFormatError(Exception):
    """Raised when the uploaded file is not a valid Excel format."""

    def __init__(self, message: str = "Invalid file format. Only .xlsx and .xls files are accepted."):
        self.message = message
        super().__init__(self.message)


class FileTooLargeError(Exception):
    """Raised when the uploaded file exceeds the maximum allowed size."""

    def __init__(self, message: str = "File exceeds maximum allowed size of 10MB."):
        self.message = message
        super().__init__(self.message)


class RowLimitExceededError(Exception):
    """Raised when the uploaded file contains more rows than allowed."""

    def __init__(self, message: str = "File exceeds maximum allowed row count of 1000."):
        self.message = message
        super().__init__(self.message)


# Required address fields that must be present in each row
_REQUIRED_FIELDS = ["street_line_1", "city", "state", "zipcode"]

# Expected column headers in the Excel file
_EXPECTED_COLUMNS = ["street_line_1", "street_line_2", "city", "state", "zipcode"]


class BulkProcessor:
    """Processes Excel files containing addresses for bulk validation.

    Validates the uploaded file format and size, reads rows, validates each
    address via AddressValidator, and returns results preserving row order.
    """

    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
    MAX_ROWS = 1000

    def __init__(self, validator: AddressValidator):
        self.validator = validator

    async def process_file(self, file: UploadFile) -> BulkValidationResult:
        """Read an Excel file, validate each row, and return ordered results.

        Args:
            file: The uploaded Excel file (UploadFile from FastAPI).

        Returns:
            A BulkValidationResult containing total_rows and ordered results.

        Raises:
            InvalidFileFormatError: If the file is not .xlsx or .xls format.
            FileTooLargeError: If the file exceeds 10MB.
            RowLimitExceededError: If the file contains more than 1000 rows.
        """
        self._validate_file_format(file)
        contents = await self._read_and_validate_size(file)
        rows = self._parse_excel(contents)
        self._validate_row_count(rows)
        results = await self._validate_rows(rows)

        return BulkValidationResult(
            total_rows=len(rows),
            results=results,
        )

    def _validate_file_format(self, file: UploadFile) -> None:
        """Check that the file has a valid Excel extension.

        Args:
            file: The uploaded file to check.

        Raises:
            InvalidFileFormatError: If the filename does not end in .xlsx or .xls.
        """
        filename = file.filename or ""
        if not filename.lower().endswith((".xlsx", ".xls")):
            raise InvalidFileFormatError()

    async def _read_and_validate_size(self, file: UploadFile) -> bytes:
        """Read the file contents and validate that it doesn't exceed the size limit.

        Args:
            file: The uploaded file to read.

        Returns:
            The raw bytes of the file.

        Raises:
            FileTooLargeError: If the file exceeds MAX_FILE_SIZE.
        """
        contents = await file.read()
        if len(contents) > self.MAX_FILE_SIZE:
            raise FileTooLargeError()
        return contents

    def _parse_excel(self, contents: bytes) -> List[dict]:
        """Parse Excel file contents into a list of row dictionaries.

        Expects the first row to contain column headers matching the expected
        address fields. Each subsequent row is converted to a dictionary keyed
        by the header values.

        Args:
            contents: The raw bytes of the Excel file.

        Returns:
            A list of dictionaries, one per data row.

        Raises:
            InvalidFileFormatError: If the file cannot be parsed as Excel.
        """
        import io

        try:
            wb = load_workbook(filename=io.BytesIO(contents), read_only=True, data_only=True)
        except Exception:
            raise InvalidFileFormatError("File could not be read as a valid Excel workbook.")

        ws = wb.active
        if ws is None:
            raise InvalidFileFormatError("Excel file contains no worksheets.")

        rows_iter = ws.iter_rows(values_only=True)

        # Read header row
        try:
            header_row = next(rows_iter)
        except StopIteration:
            # Empty file — no header, no data
            return []

        # Normalize headers to lowercase stripped strings
        headers = [
            str(cell).strip().lower() if cell is not None else ""
            for cell in header_row
        ]

        # Parse data rows
        data_rows: List[dict] = []
        for row in rows_iter:
            row_dict = {}
            for idx, cell_value in enumerate(row):
                if idx < len(headers):
                    key = headers[idx]
                    row_dict[key] = str(cell_value).strip() if cell_value is not None else ""
            data_rows.append(row_dict)

        wb.close()
        return data_rows

    def _validate_row_count(self, rows: List[dict]) -> None:
        """Validate that the number of data rows does not exceed the limit.

        Args:
            rows: The list of parsed data rows.

        Raises:
            RowLimitExceededError: If there are more than MAX_ROWS rows.
        """
        if len(rows) > self.MAX_ROWS:
            raise RowLimitExceededError()

    async def _validate_rows(self, rows: List[dict]) -> List[ValidationResult]:
        """Validate each row's address, preserving order.

        For rows with missing required fields, returns a ValidationResult with
        status "invalid_input" and a description of the missing fields.

        Args:
            rows: The list of parsed row dictionaries.

        Returns:
            A list of ValidationResult, one per row, in the same order.
        """
        results: List[ValidationResult] = []

        for row in rows:
            # Extract address fields from the row
            street_line_1 = row.get("street_line_1", "")
            street_line_2 = row.get("street_line_2", "") or None
            city = row.get("city", "")
            state = row.get("state", "")
            zipcode = row.get("zipcode", "")

            # Check for missing required fields
            missing_fields = []
            if not street_line_1:
                missing_fields.append("street_line_1")
            if not city:
                missing_fields.append("city")
            if not state:
                missing_fields.append("state")
            if not zipcode:
                missing_fields.append("zipcode")

            if missing_fields:
                # Create a StructuredAddress with what we have for the original_address
                address = StructuredAddress(
                    street_line_1=street_line_1 or "",
                    street_line_2=street_line_2,
                    city=city or "",
                    state=state or "",
                    zipcode=zipcode or "",
                )
                results.append(
                    ValidationResult(
                        original_address=address,
                        standardized_address=None,
                        status="invalid_input",
                        error_message=f"Missing required fields: {', '.join(missing_fields)}",
                    )
                )
            else:
                # Build a StructuredAddress and validate via the AddressValidator
                address = StructuredAddress(
                    street_line_1=street_line_1,
                    street_line_2=street_line_2,
                    city=city,
                    state=state,
                    zipcode=zipcode,
                )
                result = await self.validator.validate(address)
                results.append(result)

        return results
